from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def save_to_excel(data, filename="output.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    # ---------- Styles ----------
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ---------- Header ----------
    headers = ["Tag", "Text", "Link"]
    ws.append(headers)

    for col_num, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    # ---------- Data ----------
    for row in data:
        ws.append([
            row.get("tag"),
            row.get("text"),
            row.get("link")
        ])

    # ---------- Apply styling to all cells ----------
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)

    # ---------- Auto column width ----------
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 5, 50)

    # ---------- Freeze header ----------
    ws.freeze_panes = "A2"

    wb.save(filename)
    return filename